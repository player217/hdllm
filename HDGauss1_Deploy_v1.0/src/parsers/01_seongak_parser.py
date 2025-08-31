# parsers/seongak_parser.py
# 이 코드를 GUI 애플리케이션의 '로직 파일'로 사용하세요.

import os
import re
import json
import pathlib
import unicodedata
import logging
from typing import List, Dict, Any, Sequence
from concurrent.futures import ThreadPoolExecutor, as_completed

from tqdm import tqdm
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell import Cell
from langchain.docstore.document import Document

# ----------------- 전역 상수 및 정규식 -----------------
EXCEL_EXT = (".xlsx", ".xlsm")
RE_AGENDA_SHEET_NAME = re.compile(r'^\d{1,4}(-\d+)?$')

# ----------------- 유틸리티 함수 -----------------
def _long_path(p: pathlib.Path) -> str:
    s = str(p.resolve())
    return r"\\?\UNC" + s[1:] if s.startswith("\\\\") else s

def _norm_sheet_name(name: str) -> str:
    s = str(name).strip().lstrip("'")
    s = re.sub(r"[\s\u00A0]+", "", unicodedata.normalize("NFKC", s))
    return s

def _filename_meeting_no(path: pathlib.Path) -> str:
    m = re.search(r"(\d{1,4})차", path.stem)
    if not m:
        raise ValueError(f"파일명에서 회의 차수 패턴('N차')을 찾을 수 없습니다: {path.name}")
    return m.group(1)

def _cell_to_text(c: Cell) -> str:
    if c is None or c.value is None: return ""
    return str(c.value).strip()

def _join_column(ws: Worksheet, col: str, row_from: int, row_to: int) -> List[str]:
    out = [v for r in range(row_from, row_to + 1) if (v := _cell_to_text(ws[f"{col}{r}"])) ]
    return out

def _clean_text(t: str) -> str:
    t = re.sub(r'\.{2,}', ' ', t)
    t = re.sub(r'[^\w\s()\[\]\-–—.,:;/%&±→←×°]', ' ', t)
    return re.sub(r'\s+', ' ', t).strip()

def _range_to_text(ws: Worksheet, rng: str) -> str:
    lines = [" ".join(_cell_to_text(c) for c in row) for row in ws[rng]]
    return "\n".join(filter(None, lines))

# ----------------- 시트 파서 -----------------
def _parse_cover(ws: Worksheet, meta_base: Dict[str, Any]) -> Document:
    meeting_no_num = meta_base['회의 차수_번호']
    agendas: List[Dict[str, str]] = []
    for row_cells in ws.iter_rows(min_row=12, max_row=100, min_col=2, max_col=3):
        b_val, c_val = _cell_to_text(row_cells[0]), _cell_to_text(row_cells[1])
        if b_val and RE_AGENDA_SHEET_NAME.match(b_val):
            agenda_no = f"{meeting_no_num}-{b_val}" if b_val.isdigit() else b_val
            agendas.append({"안건번호": agenda_no, "안건명": c_val})
    meta = {
        "DATA_TYPE": "선각계열회의 표지", "회의 차수": f"{meeting_no_num}차",
        "회의 일시": _cell_to_text(ws["B7"]),
        "참석 인원": ", ".join(_join_column(ws, "H", 12, 50)),
        "안건 리스트": json.dumps(agendas, ensure_ascii=False),
        "path": f"{meta_base['path']}#{ws.title}",
    }
    page_content = "\n".join(f"{a['안건번호']} {a['안건명']}" for a in agendas)
    return Document(page_content=page_content, metadata=meta)

def _parse_summary(ws: Worksheet, meta_base: Dict[str, Any]) -> Document:
    meeting_no_num = meta_base['회의 차수_번호']
    summaries: List[Dict[str, Any]] = []
    for r in ws.iter_rows(min_row=5, values_only=True):
        if all(c is None for c in r): continue
        a = str(r[0] or "").strip()
        if not RE_AGENDA_SHEET_NAME.match(a): continue
        issue_no = f"{meeting_no_num}-{a}" if a.isdigit() else a
        status = ""
        status_map = {2: "적용", 3: "미적용", 4: "재검토", 5: "기타"}
        for idx, stat in status_map.items():
            if len(r) > idx and r[idx] and str(r[idx]).strip():
                status = stat; break
        summaries.append({
            "안건번호": issue_no, "안건명": str(r[1] or "").strip(),
            "적용 여부": status, "적용비고": str(r[6] or "").strip(),
            "제안부서": str(r[7] or "").strip(), "검토부서": str(r[8] or "").strip(),
        })
    meta = {
        "DATA_TYPE": "선각계열회의 회의 요약",
        "요약정보": json.dumps(summaries, ensure_ascii=False),
        "path": f"{meta_base['path']}#{ws.title}",
    }
    page_content = "\n".join(f"{s['안건번호']} {s['안건명']} ({s['적용 여부']})" for s in summaries if s.get('적용 여부'))
    return Document(page_content=page_content, metadata=meta)

# --- 안건 파서 (수정됨) ---
def _parse_agenda(ws: Worksheet, meta_cover: Dict[str, Any], summary_map: Dict[str, Dict[str, Any]], shapes_data: List[Dict] = None) -> Document:
    
    # 전달받은 텍스트 상자 데이터에서 위치에 맞는 텍스트를 가져오는 헬퍼 함수
    def get_text_from_shapes(min_col, max_col, min_row, max_row):
        if not shapes_data:
            return ""
        
        content = []
        for shape in shapes_data:
            if min_row <= shape['row'] <= max_row and min_col <= shape['column'] <= max_col:
                content.append(shape['text'])
        return "\n".join(content)

    sheet_name_norm = _norm_sheet_name(ws.title)
    meeting_no_num_from_cover = meta_cover["회의 차수"].replace("차", "")
    issue_no, carry_over = (f"{meeting_no_num_from_cover}-{sheet_name_norm}", None) if sheet_name_norm.isdigit() else (sheet_name_norm, sheet_name_norm)
    meeting_no_from_cell = _cell_to_text(ws["N2"])
    final_meeting_no_str = meeting_no_from_cell if "차" in meeting_no_from_cell else (f"{meeting_no_from_cell}차" if meeting_no_from_cell else meta_cover["회의 차수"])
    
    # 셀 데이터와 텍스트 상자 데이터를 합쳐서 각 항목 구성
    현상 = _clean_text(f'{_range_to_text(ws, "B5:F31")}\n{get_text_from_shapes(2, 6, 5, 31)}')
    개선안 = _clean_text(f'{_range_to_text(ws, "G5:K31")}\n{get_text_from_shapes(7, 11, 5, 31)}')
    결과상세 = _clean_text(f'{_range_to_text(ws, "L10:O30")}\n{get_text_from_shapes(12, 15, 10, 30)}')
    
    결과 = summary_map.get(issue_no, {}).get("적용 여부")
    
    meta = {
        "DATA_TYPE": "선각계열회의 안건", "회의 차수": final_meeting_no_str,
        "이관 차수": carry_over, "회의 일시": meta_cover["회의 일시"],
        "참석 인원": meta_cover["참석 인원"], "안건 번호": issue_no,
        "안건명": _cell_to_text(ws["C3"]), "제안부서": _cell_to_text(ws["L4"]),
        "현상": 현상, "개선안": 개선안, "결과": 결과, "결과 상세": 결과상세,
        "path": f"{meta_cover['path']}#{ws.title}",
    }
    page_content = "\n\n".join(filter(None, [현상, 개선안, 결과상세])).strip()
    return Document(page_content=page_content, metadata=meta)

# --- 워크북 로더 (수정됨) ---
def _parse_workbook(xl_path: pathlib.Path, shapes_by_sheet: dict = None) -> List[Document]:
    if shapes_by_sheet is None:
        shapes_by_sheet = {}
        
    try:
        # shapes 정보는 미리 처리했으므로, openpyxl은 read_only 모드로 빠르게 실행
        wb = load_workbook(_long_path(xl_path), data_only=True, read_only=True)
        meeting_no_num = _filename_meeting_no(xl_path)
    except Exception as e:
        logging.warning(f"파일 처리 건너뜀 '{xl_path.name}': {e}")
        return []

    docs, summary_map = [], {}
    meta_base = {"회의 차수_번호": meeting_no_num, "path": str(xl_path.resolve())}
    
    ws_cover = next((ws for ws in wb.worksheets if "표지" in _norm_sheet_name(ws.title)), None)
    if not ws_cover: return []
    
    try:
        doc_cover = _parse_cover(ws_cover, meta_base)
        docs.append(doc_cover)
        meta_cover = doc_cover.metadata
        
        ws_summary = next((ws for ws in wb.worksheets if "요약" in _norm_sheet_name(ws.title)), None)
        if ws_summary:
            doc_summary = _parse_summary(ws_summary, meta_base)
            docs.append(doc_summary)
            summary_list: Sequence[Dict] = json.loads(doc_summary.metadata.get("요약정보", "[]"))
            summary_map = {item["안건번호"]: item for item in summary_list}

        for ws in wb.worksheets:
            if RE_AGENDA_SHEET_NAME.match(_norm_sheet_name(ws.title)):
                # 현재 시트 이름에 맞는 텍스트 상자 데이터를 가져와서 전달
                sheet_shapes = shapes_by_sheet.get(ws.title, [])
                docs.append(_parse_agenda(ws, meta_cover, summary_map, shapes_data=sheet_shapes))
    except Exception as e:
        logging.error(f"'{xl_path.name}' 워크북 파싱 중 오류: {e}", exc_info=True)
    return docs

# --- 전체 로더 (주의: 이 함수는 GUI 툴에서는 직접 사용되지 않음) ---
def run_parser(folder_path: pathlib.Path) -> List[Document]:
    # 이 함수는 폴더 전체를 처리하므로, 텍스트 상자 처리를 위해서는 별도 수정이 필요합니다.
    # GUI 툴은 _parse_workbook을 직접 호출하므로 텍스트 상자 처리가 정상적으로 작동합니다.
    files = [p for p in folder_path.rglob("*.xls*") if p.suffix.lower() in EXCEL_EXT and not p.name.startswith("~$")]
    all_docs: List[Document] = []

    with ThreadPoolExecutor(max_workers=os.cpu_count() or 1) as ex:
        future_to_file = {ex.submit(_parse_workbook, f): f for f in files}
        for future in tqdm(as_completed(future_to_file), total=len(files), desc="선각계열회의록 파싱 중"):
            try:
                if part := future.result():
                    all_docs.extend(part)
            except Exception as exc:
                logging.error(f"'{future_to_file[future].name}' 처리 중 치명적 오류: {exc}", exc_info=True)
    return all_docs